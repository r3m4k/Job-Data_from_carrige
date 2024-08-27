import numpy as np
import matplotlib.pyplot as plt
import openpyxl


def rolling_mean(arr, rolling_n):
    """
    Функция для рассчёта плавающего среднего
    :param arr: исходный массив
    :param rolling_n: параметр плавающего стреднего
    :return: массив с плавающим средним длины len(arr)
    """

    rolling_arr = np.zeros(len(arr))
    for i in range(0, len(arr)):
        if i < rolling_n - 1:
            rolling_arr[i] = None
        else:
            rolling_arr[i] = np.sum(arr[i - rolling_n + 1: i + 1]) / rolling_n

    return rolling_arr


def gap(arr1, arr2):
    result = np.array([arr1[j] - arr2[int(j + rolling_n / 2)] for j in range(0, len(arr1) - rolling_n)])
    for j in range(len(result)):
        if result[j] > 300:
            result[j] = None

    return result


file = "D:/Работа/Данные с телеги/Эталон/Эталон тупика Приложение № 1 (Табл.).xlsx"
rolling_n = 50

sheet = openpyxl.load_workbook(filename=file)['проект']

min_row = 155
max_row = 871

x_axis = np.linspace(0, max_row - min_row - 1, max_row - min_row)
plan_left = np.zeros(max_row - min_row)
plan_right = np.zeros(max_row - min_row)

for i in range(min_row, max_row):
    plan_right[i - min_row] = sheet.cell(row=i, column=6).value
    try:
        plan_left[i - min_row] = plan_right[i - min_row] - sheet.cell(row=i, column=8).value
    except Exception:
        plan_left[i - min_row] = plan_right[i - min_row]

fig_right, ax_right = plt.subplots(nrows=2, ncols=1,
                                   gridspec_kw={'width_ratios': [1],
                                                'height_ratios': [3, 1]})

plt.rcParams['font.size'] = '14'
fig_right.set_figheight(9)
fig_right.set_figwidth(15)

rolling_right = rolling_mean(plan_right, rolling_n)
ax_right[0].plot(x_axis, plan_right, color='tab:blue')
ax_right[0].plot(x_axis - rolling_n / 2, rolling_right, color='tab:red')
ax_right[1].plot(np.linspace(0, max_row - min_row - 1 - rolling_n / 2, max_row - min_row - rolling_n) + rolling_n / 2,
                 gap(plan_right, rolling_right), color='tab:green')

ax_right[0].set_title('Правый план')
ax_right[0].set_xlim(-50, max_row - min_row + 50)
ax_right[0].grid()

ax_right[1].set_title('Шум')
ax_right[1].set_xlim(-50, max_row - min_row + 50)
ax_right[1].grid()

fig_right.tight_layout()


#########################################################################
"""
fig_left, ax_left = plt.subplots(nrows=2, ncols=1,
                                 gridspec_kw={'width_ratios': [1],
                                                'height_ratios': [2, 1]})
plt.rcParams['font.size'] = '14'
fig_left.set_figheight(9)
fig_left.set_figwidth(15)

rolling_left = rolling_mean(plan_left, rolling_n)
ax_left[0].plot(x_axis, plan_left, label='план левый', color='tab:cyan')
ax_left[0].plot(x_axis - rolling_n / 2, rolling_left, color='tab:purple')
ax_left[1].plot(np.linspace(0, max_row - min_row - 1 - rolling_n / 2, max_row - min_row - rolling_n) + rolling_n / 2,
                gap(plan_left, rolling_left), color='tab:olive')

ax_left[0].set_title('Левый план')
ax_left[0].set_xlim(-50, max_row - min_row + 50)
ax_left[0].grid()

ax_left[1].set_title('Шум')
ax_left[1].set_xlim(-50, max_row - min_row + 50)
ax_left[1].grid()

fig_left.tight_layout()
"""
plt.show()
